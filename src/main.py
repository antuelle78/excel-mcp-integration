import os
import logging
import re
import json
import csv
import io
from pathlib import Path
from typing import List, Optional, Dict, Any
from fastmcp import FastMCP
from openpyxl import Workbook, load_workbook
from openpyxl.styles import Font, Alignment, PatternFill, Border, Side
from openpyxl.utils import get_column_letter
from openpyxl.chart import BarChart, LineChart, PieChart, ScatterChart, Reference
from http.server import HTTPServer, SimpleHTTPRequestHandler
import threading
import urllib.parse

# Configure logging
logging.basicConfig(
    level=logging.INFO, format="%(asctime)s - %(name)s - %(levelname)s - %(message)s"
)
logger = logging.getLogger(__name__)

# Configuration
MAX_ROWS = int(os.getenv("MAX_ROWS", "10000"))
MAX_COLS = int(os.getenv("MAX_COLS", "100"))
MAX_FILENAME_LENGTH = int(os.getenv("MAX_FILENAME_LENGTH", "255"))
ALLOWED_EXTENSIONS = {".xlsx", ".xls"}
OUTPUT_DIR = os.getenv("OUTPUT_DIR", "./output")

app = FastMCP()


def validate_filename(filename: str) -> str:
    """Validate and sanitize filename for security."""
    if not filename:
        raise ValueError("Filename cannot be empty")

    if len(filename) > MAX_FILENAME_LENGTH:
        raise ValueError(f"Filename too long (max {MAX_FILENAME_LENGTH} characters)")

    # Check for dangerous characters
    dangerous_chars = ["/", "\\", "..", "<", ">", ":", "*", "?", '"', "|"]
    for char in dangerous_chars:
        if char in filename:
            raise ValueError(f"Filename contains dangerous character: {char}")

    # Ensure proper extension
    if not any(filename.lower().endswith(ext) for ext in ALLOWED_EXTENSIONS):
        filename += ".xlsx"

    # Create output directory if it doesn't exist
    output_path = Path(OUTPUT_DIR)
    output_path.mkdir(exist_ok=True)

    # Return full path
    return str(output_path / filename)


def validate_excel_data(headers: List[str], sheet_data: List[List[str]]) -> None:
    """Validate Excel data structure and size."""
    if not headers:
        raise ValueError("Headers list cannot be empty")

    if len(headers) > MAX_COLS:
        raise ValueError(f"Too many columns (max {MAX_COLS})")

    if not sheet_data:
        # Add placeholder row with empty values to satisfy Excel requirements
        sheet_data = [[""] * len(headers)]

    if len(sheet_data) > MAX_ROWS:
        raise ValueError(f"Too many rows (max {MAX_ROWS})")

    # Validate each row has correct number of columns
    expected_cols = len(headers)
    for i, row in enumerate(sheet_data):
        if len(row) != expected_cols:
            raise ValueError(
                f"Row {i + 1} has {len(row)} columns, expected {expected_cols}"
            )


def validate_excel_request(
    filename: str, headers: List[str], sheet_data: List[List[str]]
) -> Dict[str, Any]:
    """Comprehensive validation for Excel creation requests."""
    errors = []
    warnings = []

    # Validate filename
    if not filename:
        errors.append("Filename is required")
    elif len(filename) > MAX_FILENAME_LENGTH:
        errors.append(f"Filename too long (max {MAX_FILENAME_LENGTH} characters)")

    # Validate headers
    if not headers:
        errors.append("Headers array cannot be empty")
    elif len(headers) > MAX_COLS:
        errors.append(f"Too many columns (max {MAX_COLS})")
    elif any(not isinstance(h, str) for h in headers):
        errors.append("All headers must be strings")

    # Validate data
    if not sheet_data:
        errors.append("Sheet data cannot be empty")
    elif len(sheet_data) > MAX_ROWS:
        errors.append(f"Too many rows (max {MAX_ROWS})")
    else:
        # Check data consistency
        expected_cols = len(headers)
        for i, row in enumerate(sheet_data):
            if not isinstance(row, list):
                errors.append(f"Row {i + 1} is not an array")
            elif len(row) != expected_cols:
                errors.append(
                    f"Row {i + 1} has {len(row)} columns, expected {expected_cols}"
                )
            elif any(not isinstance(cell, str) for cell in row):
                warnings.append(
                    f"Row {i + 1} contains non-string values (will be converted)"
                )

    return {"valid": len(errors) == 0, "errors": errors, "warnings": warnings}


def apply_formatting(
    ws, headers: List[str], formatting: Optional[Dict[str, Any]] = None
) -> None:
    """Apply formatting to the worksheet."""
    if not formatting:
        return

    # Header formatting
    header_font = Font(bold=True)
    header_alignment = Alignment(horizontal="center")

    for col_num, _ in enumerate(headers, 1):
        cell = ws.cell(row=1, column=col_num)
        cell.font = header_font
        cell.alignment = header_alignment

    # Auto-adjust column widths
    for col_num, _ in enumerate(headers, 1):
        column_letter = get_column_letter(col_num)
        max_length = 0
        for row_num in range(1, len(ws[column_letter]) + 1):
            cell_value = str(ws.cell(row=row_num, column=col_num).value or "")
            max_length = max(max_length, len(cell_value))
        ws.column_dimensions[column_letter].width = min(max_length + 2, 50)


def parse_cell_range(cell_range: str) -> tuple:
    """
    Parse Excel cell range notation (e.g., 'A1:C10') into row/column indices.

    Args:
        cell_range: Excel cell range in format 'A1:C10'

    Returns:
        Tuple of (start_row, end_row, start_col, end_col)

    Raises:
        ValueError: If cell_range format is invalid
    """
    try:
        if ":" not in cell_range:
            raise ValueError("Cell range must contain ':' separator")

        parts = cell_range.split(":")
        if len(parts) != 2:
            raise ValueError("Cell range must have exactly 2 parts (start:end)")

        start_cell = parts[0].strip().upper()
        end_cell = parts[1].strip().upper()

        # Parse start cell (e.g., 'A1')
        start_col_match = re.match(r"([A-Z]+)(\d+)", start_cell)
        if not start_col_match:
            raise ValueError(f"Invalid start cell: {start_cell}")
        start_col_letters = start_col_match.group(1)
        start_row = int(start_col_match.group(2))

        # Convert column letters to number (A=1, B=2, ..., Z=26, AA=27, etc.)
        start_col = 0
        for char in start_col_letters:
            start_col = start_col * 26 + (ord(char) - ord("A") + 1)

        # Parse end cell (e.g., 'C10')
        end_col_match = re.match(r"([A-Z]+)(\d+)", end_cell)
        if not end_col_match:
            raise ValueError(f"Invalid end cell: {end_cell}")
        end_col_letters = end_col_match.group(1)
        end_row = int(end_col_match.group(2))

        # Convert column letters to number
        end_col = 0
        for char in end_col_letters:
            end_col = end_col * 26 + (ord(char) - ord("A") + 1)

        return start_row, end_row, start_col, end_col
    except Exception as e:
        raise ValueError(f"Failed to parse cell range '{cell_range}': {str(e)}")


@app.resource(uri="mcp://resources/system_prompt")
def system_prompt() -> str:
    """
    Returns the system prompt for the Exel MCP server.
    """
    try:
        with open("../docs/system_prompt.txt", "r") as f:
            return f.read()
    except FileNotFoundError:
        logger.error("system_prompt.txt not found")
        return "Excel file creation assistant"


@app.tool()
def create_excel_file(
    filename: str,
    headers: List[str],
    sheet_data: List[List[Any]],
    sheet_name: str = "Sheet1",
    formatting: Optional[Dict[str, Any]] = None,
) -> str:
    """
    Creates an Excel file with the given data.

    Args:
        filename: Name of the Excel file to create
        headers: List of column headers
        sheet_data: 2D list of data rows
        sheet_name: Name of the worksheet (default: "Sheet1")
        formatting: Optional formatting options

    Returns:
        Success message with file path
    """
    try:
        logger.info(f"Creating Excel file: {filename}")

        # Validate inputs
        safe_filename = validate_filename(filename)
        validate_excel_data(headers, sheet_data)

        # Create workbook
        wb = Workbook()
        ws = wb.active
        if ws is None:
            ws = wb.create_sheet(sheet_name)
        else:
            ws.title = sheet_name

        # Add headers
        ws.append(headers)

        # Add data rows
        for row in sheet_data:
            ws.append(row)

        # Apply formatting
        apply_formatting(ws, headers, formatting)

        # Save file
        wb.save(safe_filename)
        logger.info(f"Successfully created Excel file: {safe_filename}")

        return f"Successfully created Excel file: {safe_filename}"

    except ValueError as e:
        error_msg = f"Validation error: {str(e)}"
        logger.error(error_msg)
        raise ValueError(error_msg)
    except Exception as e:
        error_msg = f"Failed to create Excel file: {str(e)}"
        logger.error(error_msg)
        raise Exception(error_msg)


@app.tool()
def get_excel_info(filename: str) -> Dict[str, Any]:
    """
    Get information about an existing Excel file.

    Args:
        filename: Name of the Excel file to analyze

    Returns:
        Dictionary with file information
    """
    try:
        safe_filename = validate_filename(filename)

        if not Path(safe_filename).exists():
            raise FileNotFoundError(f"File not found: {safe_filename}")

        # Load the actual workbook to get file information
        wb = load_workbook(safe_filename)

        # Get sheet information
        sheet_info = {}
        for sheet_name in wb.sheetnames:
            ws = wb[sheet_name]
            sheet_info[sheet_name] = {
                "dimensions": ws.dimensions,
                "max_row": ws.max_row,
                "max_column": ws.max_column,
            }

        # Get file statistics
        file_size = Path(safe_filename).stat().st_size

        result = {
            "filename": safe_filename,
            "exists": True,
            "size": file_size,
            "size_kb": round(file_size / 1024, 2),
            "sheet_count": len(wb.sheetnames),
            "sheets": wb.sheetnames,
            "active_sheet": wb.active.title if wb.active else None,
            "sheet_info": sheet_info,
        }

        wb.close()
        return result

    except Exception as e:
        error_msg = f"Failed to get Excel info: {str(e)}"
        logger.error(error_msg)
        return {"error": error_msg}


@app.tool()
def create_excel_chart(
    filename: str,
    chart_type: str,
    data_range: str,
    title: Optional[str] = None,
    sheet_name: Optional[str] = None,
) -> str:
    """
    Add charts and graphs to existing Excel files.

    Args:
        filename: Target Excel file to add chart to
        chart_type: Type of chart to create (bar, line, pie, scatter, area)
        data_range: Cell range for chart data (e.g., 'A1:C10')
        title: Chart title (optional)
        sheet_name: Worksheet name (optional, defaults to first sheet)

    Returns:
        Success message with chart details
    """
    try:
        # Set default title if not provided
        if title is None:
            title = f"{chart_type.title()} Chart"

        # Validate inputs
        if not filename or not chart_type or not data_range:
            raise ValueError("filename, chart_type, and data_range are required")

        safe_filename = validate_filename(filename)

        if not Path(safe_filename).exists():
            raise FileNotFoundError(f"Excel file not found: {safe_filename}")

        # Load existing workbook
        wb = load_workbook(safe_filename)

        # Select worksheet
        if sheet_name and sheet_name in wb.sheetnames:
            ws = wb[sheet_name]
        else:
            ws = wb.active

        if ws is None:
            raise ValueError("Worksheet not found")

        # Create chart based on type
        if chart_type == "bar":
            chart = BarChart()
        elif chart_type == "line":
            chart = LineChart()
        elif chart_type == "pie":
            chart = PieChart()
        elif chart_type == "scatter":
            chart = ScatterChart()
        else:
            chart = BarChart()  # Default to bar chart

        # Set data range
        try:
            # Parse the data_range parameter to get specific cell range
            start_row, end_row, start_col, end_col = parse_cell_range(data_range)
            data = Reference(
                ws,
                min_col=start_col,
                min_row=start_row,
                max_col=end_col,
                max_row=end_row,
            )
            chart.add_data(data, titles_from_data=True)
        except Exception as e:
            raise ValueError(f"Invalid data range '{data_range}': {str(e)}")

        # Set chart title
        chart.title = title

        # Add chart to worksheet
        ws.add_chart(chart)

        # Save workbook
        wb.save(safe_filename)
        logger.info(f"Successfully added {chart_type} chart to {safe_filename}")

        return f"Successfully added {chart_type} chart '{title}' to {safe_filename}"

    except Exception as e:
        error_msg = f"Failed to create chart: {str(e)}"
        logger.error(error_msg)
        raise Exception(error_msg)


@app.tool()
def format_excel_cells(
    filename: str,
    cell_range: str,
    formatting: Dict[str, Any],
    sheet_name: Optional[str] = None,
) -> str:
    """
    Apply formatting to Excel cells including colors, borders, fonts, and styles.

    Args:
        filename: Target Excel file
        cell_range: Cell range in A1:B5 format (e.g., 'A1:C10')
        formatting: Formatting options to apply
        sheet_name: Worksheet name (optional, defaults to first sheet)

    Returns:
        Success message with formatting details
    """
    try:
        # Validate inputs
        if not filename or not cell_range or not formatting:
            raise ValueError("filename, cell_range, and formatting are required")

        safe_filename = validate_filename(filename)

        if not Path(safe_filename).exists():
            raise FileNotFoundError(f"Excel file not found: {safe_filename}")

        # Load existing workbook
        wb = load_workbook(safe_filename)

        # Select worksheet
        if sheet_name and sheet_name in wb.sheetnames:
            ws = wb[sheet_name]
        else:
            ws = wb.active

        if ws is None:
            raise ValueError("Worksheet not found")

        # Parse cell range and apply formatting
        try:
            # Simple approach: iterate through range manually
            # Parse the cell_range using the helper function
            start_row, end_row, start_col, end_col = parse_cell_range(cell_range)

            # Apply formatting to all cells in the range
            for row_idx in range(start_row, end_row + 1):
                for col_idx in range(start_col, end_col + 1):
                    cell = ws.cell(row=row_idx, column=col_idx)

                    # Font formatting
                    font_kwargs = {}
                    if formatting.get("bold") is not None:
                        font_kwargs["bold"] = formatting["bold"]
                    if formatting.get("italic") is not None:
                        font_kwargs["italic"] = formatting["italic"]
                    if formatting.get("underline") is not None:
                        font_kwargs["underline"] = formatting["underline"]
                    if formatting.get("font_size") is not None:
                        font_kwargs["size"] = formatting["font_size"]
                    if formatting.get("font_color"):
                        font_kwargs["color"] = formatting["font_color"]

                    if font_kwargs:
                        cell.font = Font(**font_kwargs)

                    # Fill formatting
                    if formatting.get("background_color"):
                        cell.fill = PatternFill(
                            start_color=formatting["background_color"],
                            end_color=formatting["background_color"],
                            fill_type="solid",
                        )

                    # Alignment
                    if formatting.get("alignment"):
                        cell.alignment = Alignment(horizontal=formatting["alignment"])

                    # Border
                    if formatting.get("border"):
                        border_color = formatting.get("border_color", "000000")
                        thin_border = Border(
                            left=Side(style="thin", color=border_color),
                            right=Side(style="thin", color=border_color),
                            top=Side(style="thin", color=border_color),
                            bottom=Side(style="thin", color=border_color),
                        )
                        cell.border = thin_border

        except Exception as e:
            raise ValueError(f"Invalid cell range '{cell_range}': {str(e)}")

        # Save workbook
        wb.save(safe_filename)
        logger.info(
            f"Successfully applied formatting to {cell_range} in {safe_filename}"
        )

        return f"Successfully applied formatting to {cell_range} in {safe_filename}"

    except Exception as e:
        error_msg = f"Failed to format cells: {str(e)}"
        logger.error(error_msg)
        raise Exception(error_msg)


@app.tool()
def import_csv_to_excel(
    csv_file: str,
    excel_file: str,
    delimiter: str = ",",
    has_headers: bool = True,
    sheet_name: str = "Sheet1",
) -> str:
    """
    Convert CSV files to Excel format with proper formatting and structure.

    Args:
        csv_file: Source CSV file path or content
        excel_file: Target Excel filename
        delimiter: CSV delimiter character (default: ',')
        has_headers: Whether CSV has header row (default: true)
        sheet_name: Worksheet name (optional, defaults to 'Sheet1')

    Returns:
        Success message with file path
    """
    try:
        # Validate inputs
        if not csv_file or not excel_file:
            raise ValueError("csv_file and excel_file are required")

        safe_excel_file = validate_filename(excel_file)

        # Read CSV data
        if os.path.exists(csv_file):
            # Read from file
            with open(csv_file, "r", encoding="utf-8") as f:
                csv_reader = csv.reader(f, delimiter=delimiter)
                rows = list(csv_reader)
        else:
            # Treat as CSV content
            csv_reader = csv.reader(csv_file.splitlines(), delimiter=delimiter)
            rows = list(csv_reader)

        if not rows:
            raise ValueError("CSV file is empty")

        # Create workbook
        wb = Workbook()
        ws = wb.active
        if ws is not None:
            ws.title = sheet_name

        # Write data to worksheet
        headers = None
        data_rows = []

        if has_headers and rows:
            headers = rows[0]
            data_rows = rows[1:]
        else:
            headers = [f"Column {i + 1}" for i in range(len(rows[0]))]
            data_rows = rows

        # Add headers
        if ws is not None:
            ws.append(headers)

        # Add data rows
        for row in data_rows:
            if ws is not None:
                ws.append(row)

        # Apply basic formatting
        apply_formatting(ws, headers, {"auto_width": True, "header_bold": True})

        # Save workbook
        wb.save(safe_excel_file)
        logger.info(f"Successfully converted CSV to Excel: {safe_excel_file}")

        return f"Successfully converted CSV to Excel: {safe_excel_file}"

    except Exception as e:
        error_msg = f"Failed to import CSV to Excel: {str(e)}"
        logger.error(error_msg)
        raise Exception(error_msg)


@app.tool()
def export_excel_to_csv(
    excel_file: str,
    csv_file: str,
    sheet_name: Optional[str] = None,
    delimiter: str = ",",
    include_headers: bool = True,
) -> str:
    """
    Export Excel worksheets to CSV format.

    Args:
        excel_file: Source Excel file
        csv_file: Target CSV filename
        sheet_name: Worksheet name to export (optional, defaults to first sheet)
        delimiter: CSV delimiter character (default: ',')
        include_headers: Whether to include headers in CSV (default: true)

    Returns:
        Success message with CSV file path
    """
    try:
        # Validate inputs
        if not excel_file or not csv_file:
            raise ValueError("excel_file and csv_file are required")

        safe_excel_file = validate_filename(excel_file)

        if not Path(safe_excel_file).exists():
            raise FileNotFoundError(f"Excel file not found: {safe_excel_file}")

        # Load workbook
        wb = load_workbook(safe_excel_file)

        # Select worksheet
        if sheet_name and sheet_name in wb.sheetnames:
            ws = wb[sheet_name]
        else:
            ws = wb.active

        # Read data from worksheet
        data = []
        if ws is not None:
            for row in ws.iter_rows(values_only=True):
                # Convert None to empty string
                row_data = [cell if cell is not None else "" for cell in row]
                data.append(row_data)

        if not data:
            raise ValueError("Worksheet is empty")

        # Create output directory if needed
        output_dir = Path(csv_file).parent
        output_dir.mkdir(parents=True, exist_ok=True)

        # Write to CSV
        with open(csv_file, "w", newline="", encoding="utf-8") as f:
            csv_writer = csv.writer(f, delimiter=delimiter)

            # Write headers if requested
            if include_headers and data:
                csv_writer.writerow(data[0])
                data = data[1:]

            # Write data rows
            csv_writer.writerows(data)

        logger.info(f"Successfully exported Excel to CSV: {csv_file}")

        return f"Successfully exported Excel to CSV: {csv_file}"

    except Exception as e:
        error_msg = f"Failed to export Excel to CSV: {str(e)}"
        logger.error(error_msg)
        raise Exception(error_msg)


class FileHandler(SimpleHTTPRequestHandler):
    """Custom handler to serve files from output directory."""

    def __init__(self, *args, **kwargs) -> None:
        """Initialize the file handler for serving files from output directory."""
        # Don't set directory here, we'll handle paths manually
        super().__init__(*args, **kwargs)

    def do_GET(self) -> None:
        """Handle HTTP GET requests for file downloads."""
        # Parse the path
        parsed_path = urllib.parse.urlparse(self.path)
        file_path = parsed_path.path.lstrip("/")

        # Check if this is a file request
        if file_path.startswith("files/"):
            filename = file_path[6:]  # Remove 'files/' prefix
            file_full_path = Path(OUTPUT_DIR) / filename

            # Security check - ensure file is within output directory
            try:
                file_full_path.resolve().relative_to(Path(OUTPUT_DIR).resolve())
            except ValueError:
                self.send_error(403, "Access denied")
                return

            # Check if file exists
            if file_full_path.exists() and file_full_path.is_file():
                # Serve the file
                self.send_response(200)
                content_type = self.guess_type(str(file_full_path))
                self.send_header("Content-Type", content_type)
                self.send_header(
                    "Content-Disposition", f'attachment; filename="{filename}"'
                )
                self.send_header("Content-Length", str(file_full_path.stat().st_size))
                self.end_headers()

                with open(file_full_path, "rb") as f:
                    self.wfile.write(f.read())
            else:
                self.send_error(404, f"File not found: {filename}")
        else:
            # Default behavior for other paths
            self.send_error(404, "Not found")

    def do_HEAD(self) -> None:
        """Handle HTTP HEAD requests for file existence checks."""
        parsed_path = urllib.parse.urlparse(self.path)
        file_path = parsed_path.path.lstrip("/")

        if file_path.startswith("files/"):
            filename = file_path[6:]  # Remove 'files/' prefix
            file_full_path = Path(OUTPUT_DIR) / filename

            # Security check
            try:
                file_full_path.resolve().relative_to(Path(OUTPUT_DIR).resolve())
            except ValueError:
                self.send_error(403, "Access denied")
                return

            if file_full_path.exists() and file_full_path.is_file():
                self.send_response(200)
                content_type = self.guess_type(str(file_full_path))
                self.send_header("Content-Type", content_type)
                self.send_header(
                    "Content-Disposition", f'attachment; filename="{filename}"'
                )
                self.send_header("Content-Length", str(file_full_path.stat().st_size))
                self.end_headers()
            else:
                self.send_error(404, f"File not found: {filename}")
        else:
            self.send_error(404, "Not found")


def start_file_server():
    """Start a simple HTTP server for file downloads."""
    file_server_port = int(os.getenv("FILE_SERVER_PORT", "8001"))
    server = HTTPServer(("0.0.0.0", file_server_port), FileHandler)
    logger.info(f"File server started on port {file_server_port}")
    server.serve_forever()


if __name__ == "__main__":
    host = os.getenv("HOST", "0.0.0.0")
    port = int(os.getenv("PORT", "8000"))

    logger.info(f"Starting Exel MCP server on {host}:{port}")
    logger.info(f"Output directory: {OUTPUT_DIR}")
    logger.info(f"Max rows: {MAX_ROWS}, Max columns: {MAX_COLS}")

    # Start file server in a separate thread
    file_server_thread = threading.Thread(target=start_file_server, daemon=True)
    file_server_thread.start()

    # Start MCP server
    app.run(transport="http", host=host, port=port)
